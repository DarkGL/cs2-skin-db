"""
CS2 Skin Mapping Generator
Fetches items_game.txt + csgo_english.txt from SteamDatabase GitHub,
then generates skin mapping files in data/.

Output files:
  skin-market-mapping.json   - defindex, paint_index, market_hash_name, wear range
  skin-market-mapping.xlsx   - Same data as a styled Excel spreadsheet
"""

import re
import json
import sys
import urllib.request
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_DIR     = SCRIPT_DIR / "data"
CACHE_DIR   = SCRIPT_DIR / ".cache"

ITEMS_GAME_URL  = "https://raw.githubusercontent.com/SteamDatabase/GameTracking-CS2/refs/heads/master/game/csgo/pak01_dir/scripts/items/items_game.txt"
ENGLISH_TXT_URL = "https://raw.githubusercontent.com/SteamDatabase/GameTracking-CS2/refs/heads/master/game/csgo/pak01_dir/resource/csgo_english.txt"

OUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# ── Helpers ────────────────────────────────────────────────────────────────────
def _decode_vdf(data: bytes) -> str:
    """
    Decode raw bytes from a Valve VDF file.
    Steam files are usually UTF-16 LE (with or without BOM) or UTF-8.
    Detects encoding by checking for BOM or null-byte pattern.
    """
    # Explicit BOM check
    if data[:2] == b"\xff\xfe":
        return data[2:].decode("utf-16-le")
    if data[:2] == b"\xfe\xff":
        return data[2:].decode("utf-16-be")
    if data[:3] == b"\xef\xbb\xbf":
        return data[3:].decode("utf-8")

    # Heuristic: if >20% of bytes in the first 512 are null → UTF-16 LE
    sample = data[:512]
    null_count = sample.count(b"\x00")
    if null_count > len(sample) * 0.2:
        return data.decode("utf-16-le", errors="replace")

    # Default: UTF-8
    return data.decode("utf-8", errors="replace")


def fetch(url: str, cache_name: str, force: bool = False) -> list[str]:
    """Download url (or use cache), decode, return lines."""
    cache_path = CACHE_DIR / cache_name
    if not force and cache_path.exists() and cache_path.stat().st_size > 1000:
        print(f"  [cache] {cache_name}")
        return cache_path.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    print(f"  [fetch] {url}")
    with urllib.request.urlopen(url, timeout=60) as r:
        data = r.read()
    text = _decode_vdf(data)
    # Normalise line endings → \n
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    cache_path.write_text(text, encoding="utf-8")
    lines = text.splitlines(keepends=True)
    print(f"  -> {len(lines):,} lines, {len(data)//1024}K raw")
    return lines


# ── VDF parse helpers ──────────────────────────────────────────────────────────
_PAIR_RE   = re.compile(r'"([^"]*)"')
_LOOT_RE   = re.compile(r'^\s*"\[([^\]]+)\](weapon_[^"]+)"')


def _parse_vdf(lines: list[str]) -> tuple[dict, dict, dict, set]:
    """
    Single-pass VDF parser.
    Returns:
        paint_kits  - pk_id (str) -> {name, wear_min, wear_max, description_tag}
        items_data  - defidx (int) -> {name, item_name}
        loot_pairs  - set of (pk_name, weapon_class)
    """
    paint_kits: dict  = {}
    items_data: dict  = {}
    loot_pairs: set   = set()

    depth          = 0
    ctx_stack      = []
    cur_key        = None

    in_pk          = False
    pk_depth       = None
    cur_pk_id      = None
    cur_pk_data    = {}

    in_item        = False
    item_depth     = None
    cur_item_id    = None
    cur_item_data  = {}

    for line in lines:
        s = line.strip()

        if s == "{":
            depth += 1
            ctx_stack.append(cur_key or "__")
            parent = ctx_stack[-2] if len(ctx_stack) >= 2 else None

            if parent == "paint_kits" and cur_pk_id is not None:
                try:
                    int(cur_pk_id)
                    in_pk      = True
                    pk_depth   = depth
                    cur_pk_data = {}
                except ValueError:
                    pass

            if parent == "items" and cur_item_id is not None:
                in_item     = True
                item_depth  = depth
                cur_item_data = {}

            cur_key = None
            continue

        if s == "}":
            if in_pk and depth == pk_depth:
                if cur_pk_id and "name" in cur_pk_data:
                    paint_kits[cur_pk_id] = {
                        "name":            cur_pk_data["name"],
                        "wear_min":        float(cur_pk_data.get("wear_remap_min", "0")),
                        "wear_max":        float(cur_pk_data.get("wear_remap_max", "1")),
                        "description_tag": cur_pk_data.get("description_tag", ""),
                    }
                in_pk      = False
                cur_pk_id  = None

            if in_item and depth == item_depth:
                if cur_item_id and "name" in cur_item_data:
                    try:
                        items_data[int(cur_item_id)] = cur_item_data.copy()
                    except ValueError:
                        pass
                in_item    = False
                cur_item_id = None

            if ctx_stack:
                ctx_stack.pop()
            depth -= 1
            cur_key = None
            continue

        parts = _PAIR_RE.findall(s)
        if not parts:
            continue

        if len(parts) == 1:
            cur_key = parts[0]
            top     = ctx_stack[-1] if ctx_stack else None
            if top == "paint_kits":
                cur_pk_id = parts[0]
            if top == "items":
                try:
                    int(parts[0])
                    cur_item_id = parts[0]
                except ValueError:
                    cur_item_id = None

        elif len(parts) == 2:
            key, val = parts
            if in_pk and key in ("name", "wear_remap_min", "wear_remap_max", "description_tag"):
                cur_pk_data[key] = val
            if in_item and key in ("name", "item_name"):
                cur_item_data[key] = val

            m = _LOOT_RE.match(line)
            if m:
                loot_pairs.add((m.group(1), m.group(2)))

    return paint_kits, items_data, loot_pairs


def _parse_locale(lines: list[str]) -> dict[str, str]:
    locale: dict = {}
    for line in lines:
        parts = _PAIR_RE.findall(line)
        if len(parts) == 2:
            locale[parts[0].lower()] = parts[1]
    return locale


def loc(locale: dict, key: str) -> str:
    k = key.lstrip("#").lower()
    return locale.get(k, key)


# ── Static weapon name map (class -> display name) ────────────────────────────
WEAPON_NAMES: dict[str, str] = {
    "weapon_deagle":           "Desert Eagle",
    "weapon_elite":            "Dual Berettas",
    "weapon_fiveseven":        "Five-SeveN",
    "weapon_glock":            "Glock-18",
    "weapon_ak47":             "AK-47",
    "weapon_aug":              "AUG",
    "weapon_awp":              "AWP",
    "weapon_famas":            "FAMAS",
    "weapon_g3sg1":            "G3SG1",
    "weapon_galilar":          "Galil AR",
    "weapon_m249":             "M249",
    "weapon_m4a1":             "M4A4",
    "weapon_mac10":            "MAC-10",
    "weapon_p90":              "P90",
    "weapon_mp5sd":            "MP5-SD",
    "weapon_ump45":            "UMP-45",
    "weapon_xm1014":           "XM1014",
    "weapon_bizon":            "PP-Bizon",
    "weapon_mag7":             "MAG-7",
    "weapon_negev":            "Negev",
    "weapon_sawedoff":         "Sawed-Off",
    "weapon_tec9":             "Tec-9",
    "weapon_hkp2000":          "P2000",
    "weapon_mp7":              "MP7",
    "weapon_mp9":              "MP9",
    "weapon_nova":             "Nova",
    "weapon_p250":             "P250",
    "weapon_scar20":           "SCAR-20",
    "weapon_sg556":            "SG 553",
    "weapon_ssg08":            "SSG 08",
    "weapon_m4a1_silencer":    "M4A1-S",
    "weapon_usp_silencer":     "USP-S",
    "weapon_cz75a":            "CZ75-Auto",
    "weapon_revolver":         "R8 Revolver",
    # Knives
    "weapon_bayonet":              "Bayonet",
    "weapon_knife_css":            "Classic Knife",
    "weapon_knife_flip":           "Flip Knife",
    "weapon_knife_gut":            "Gut Knife",
    "weapon_knife_karambit":       "Karambit",
    "weapon_knife_m9_bayonet":     "M9 Bayonet",
    "weapon_knife_tactical":       "Huntsman Knife",
    "weapon_knife_falchion":       "Falchion Knife",
    "weapon_knife_survival_bowie": "Bowie Knife",
    "weapon_knife_butterfly":      "Butterfly Knife",
    "weapon_knife_push":           "Shadow Daggers",
    "weapon_knife_cord":           "Paracord Knife",
    "weapon_knife_canis":          "Survival Knife",
    "weapon_knife_ursus":          "Ursus Knife",
    "weapon_knife_gypsy_jackknife":"Navaja Knife",
    "weapon_knife_outdoor":        "Nomad Knife",
    "weapon_knife_stiletto":       "Stiletto Knife",
    "weapon_knife_widowmaker":     "Talon Knife",
    "weapon_knife_skeleton":       "Skeleton Knife",
    "weapon_knife_kukri":          "Kukri Knife",
}

KNIFE_CLASSES: set[str] = {c for c in WEAPON_NAMES if "knife" in c or c == "weapon_bayonet"}
GLOVE_PAINT_KIT_THRESHOLD = 10006  # paint_index >= this are glove kits


# ── Main ───────────────────────────────────────────────────────────────────────
def main(force_download: bool = False):
    print("=" * 60)
    print("CS2 Skin Mapping Generator")
    print("=" * 60)

    # 1. Fetch source files
    print("\n[1] Fetching source files...")
    items_lines   = fetch(ITEMS_GAME_URL,  "items_game.txt",   force=force_download)
    english_lines = fetch(ENGLISH_TXT_URL, "csgo_english.txt", force=force_download)

    # 2. Parse
    print("\n[2] Parsing...")
    paint_kits, items_data, loot_pairs = _parse_vdf(items_lines)
    locale = _parse_locale(english_lines)
    print(f"  Paint kits : {len(paint_kits)}")
    print(f"  Item defs  : {len(items_data)}")
    print(f"  Loot pairs : {len(loot_pairs)}")
    print(f"  Locale keys: {len(locale)}")

    # 3. Build lookup tables
    print("\n[3] Building lookup tables...")

    # defindex -> weapon class
    defidx_to_class: dict[int, str] = {
        did: d["name"] for did, d in items_data.items() if d.get("name", "").startswith("weapon_")
    }
    class_to_defidx: dict[str, int] = {cls: did for did, cls in defidx_to_class.items()}

    # Enrich WEAPON_NAMES from locale
    weapon_names = dict(WEAPON_NAMES)
    for did, d in items_data.items():
        cls  = d.get("name", "")
        ikey = d.get("item_name", "")
        if cls and ikey and cls not in weapon_names:
            display = loc(locale, ikey)
            if display and display != ikey:
                weapon_names[cls] = display

    # paint_kit_name -> pk_id
    pkname_to_id: dict[str, str] = {v["name"]: k for k, v in paint_kits.items()}

    # Paint kits used by weapons (i.e. in loot pairs)
    weapon_pk_names: set[str] = {pk for pk, _ in loot_pairs}

    # Knife defindexes
    knife_defidx: dict[int, tuple[str, str]] = {}  # defidx -> (class, display)
    for did, cls in defidx_to_class.items():
        if cls in KNIFE_CLASSES:
            knife_defidx[did] = (cls, weapon_names.get(cls, cls))

    # Knife paint kits = paint kits NOT in any weapon loot pair, pk_id < threshold
    knife_pks: dict[str, dict] = {
        pk_id: pk_data
        for pk_id, pk_data in paint_kits.items()
        if pk_data["name"] not in weapon_pk_names
        and pk_data["name"] not in ("default", "workshop_default")
        and int(pk_id) < GLOVE_PAINT_KIT_THRESHOLD
    }

    print(f"  Weapon defindexes : {len(defidx_to_class)}")
    print(f"  Knife defindexes  : {len(knife_defidx)}")
    print(f"  Knife paint kits  : {len(knife_pks)}")

    # 4. Build skin list (weapon skins)
    print("\n[4] Building weapon skin list...")
    weapon_skins: list[dict] = []

    for pk_name, weapon_class in sorted(loot_pairs):
        defidx = class_to_defidx.get(weapon_class)
        pk_id  = pkname_to_id.get(pk_name)
        if defidx is None or pk_id is None:
            continue
        pk_data       = paint_kits[pk_id]
        weapon_display = weapon_names.get(weapon_class, weapon_class)
        skin_display   = loc(locale, pk_data["description_tag"]) if pk_data["description_tag"] else pk_name
        is_knife       = weapon_class in KNIFE_CLASSES
        prefix         = "\u2605 " if is_knife else ""
        weapon_skins.append({
            "defindex":          defidx,
            "weapon_class":      weapon_class,
            "weapon":            weapon_display,
            "paint_index":       int(pk_id),
            "paint_kit_name":    pk_name,
            "skin":              skin_display,
            "market_hash_name":  f"{prefix}{weapon_display} | {skin_display}",
            "wear_min":          pk_data["wear_min"],
            "wear_max":          pk_data["wear_max"],
        })

    weapon_skins.sort(key=lambda x: (x["defindex"], x["paint_index"]))
    print(f"  Weapon skins: {len(weapon_skins)}")

    # 5. Build knife skin list (all knife types × all knife paint kits)
    print("\n[5] Building knife skin list...")
    knife_skins: list[dict] = []

    for pk_id, pk_data in sorted(knife_pks.items(), key=lambda x: int(x[0])):
        skin_display = loc(locale, pk_data["description_tag"]) if pk_data["description_tag"] else pk_data["name"]
        for defidx, (cls, knife_display) in sorted(knife_defidx.items()):
            knife_skins.append({
                "defindex":         defidx,
                "weapon_class":     cls,
                "weapon":           knife_display,
                "paint_index":      int(pk_id),
                "paint_kit_name":   pk_data["name"],
                "skin":             skin_display,
                "market_hash_name": f"\u2605 {knife_display} | {skin_display}",
                "wear_min":         pk_data["wear_min"],
                "wear_max":         pk_data["wear_max"],
            })

    knife_skins.sort(key=lambda x: (x["defindex"], x["paint_index"]))
    print(f"  Knife skins: {len(knife_skins)}")

    all_skins = weapon_skins + knife_skins
    all_skins.sort(key=lambda x: (x["defindex"], x["paint_index"]))

    # 6. Write files
    print("\n[6] Writing output files...")

    def write_json(path: Path, data):
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  {path.name} ({path.stat().st_size // 1024}K)")

    # skin-market-mapping.json (all skins, market hash names)
    market_rows = [
        {
            "defindex":         s["defindex"],
            "paint_index":      s["paint_index"],
            "market_hash_name": s["market_hash_name"],
            "weapon":           s["weapon"],
            "skin":             s["skin"],
            "paint_kit_name":   s["paint_kit_name"],
            "wear_min":         s["wear_min"],
            "wear_max":         s["wear_max"],
        }
        for s in all_skins
    ]
    write_json(OUT_DIR / "skin-market-mapping.json", market_rows)

    # skin-market-mapping.xlsx
    _write_excel(OUT_DIR / "skin-market-mapping.xlsx", market_rows)

    # 7. Verification
    print("\n[7] Verification samples:")
    _verify(all_skins)

    print("\nDone!")


def _write_typescript(path: Path, weapon_skins, knife_skins,
                      defidx_to_class, weapon_names, paint_kits, knife_defidx, knife_pks):
    lines = [
        "// Auto-generated by generate_skin_mappings.py — DO NOT EDIT",
        "",
        "// def_index -> weapon info",
        "export const WEAPON_DEFINDEX: Record<number, { class: string; name: string }> = {",
    ]
    for did, cls in sorted(defidx_to_class.items()):
        if cls.startswith("weapon_"):
            name = weapon_names.get(cls, cls).replace('"', '\\"')
            lines.append(f'  {did}: {{ class: "{cls}", name: "{name}" }},')
    lines += ["}", ""]

    lines += [
        "// paint_index -> { name, wear_min, wear_max }",
        "export const PAINT_KIT: Record<number, { name: string; wear_min: number; wear_max: number }> = {",
    ]
    for pk_id, pk in sorted(paint_kits.items(), key=lambda x: int(x[0])):
        lines.append(f'  {pk_id}: {{ name: "{pk["name"]}", wear_min: {pk["wear_min"]}, wear_max: {pk["wear_max"]} }},')
    lines += ["}", ""]

    lines += [
        "// Knife def_indexes",
        "export const KNIFE_DEFINDEX: Record<number, { class: string; name: string }> = {",
    ]
    for did, (cls, name) in sorted(knife_defidx.items()):
        lines.append(f'  {did}: {{ class: "{cls}", name: "{name}" }},')
    lines += ["}", ""]

    lines += [
        "// Knife paint kits — apply to ALL knife types",
        "export const KNIFE_PAINT_KIT: Record<number, { name: string; wear_min: number; wear_max: number }> = {",
    ]
    for pk_id, pk in sorted(knife_pks.items(), key=lambda x: int(x[0])):
        lines.append(f'  {pk_id}: {{ name: "{pk["name"]}", wear_min: {pk["wear_min"]}, wear_max: {pk["wear_max"]} }},')
    lines += ["}", ""]

    lines += [
        "// [defindex, paint_index, wear_min, wear_max, weapon_name, paint_kit_name]",
        "export const SKIN_MAPPING: [number, number, number, number, string, string][] = [",
    ]
    for s in weapon_skins:
        wn = s["weapon"].replace('"', '\\"')
        pk = s["paint_kit_name"].replace('"', '\\"')
        lines.append(f'  [{s["defindex"]}, {s["paint_index"]}, {s["wear_min"]}, {s["wear_max"]}, "{wn}", "{pk}"],')
    lines += ["]", ""]

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  {path.name} ({path.stat().st_size // 1024}K)")


def _write_excel(path: Path, rows: list[dict]):
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skin Market Mapping"

    headers = ["defindex", "paint_index", "market_hash_name", "weapon",
               "skin_name", "paint_kit_name", "wear_min", "wear_max"]
    ws.append(headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["defindex"], r["paint_index"], r["market_hash_name"],
            r["weapon"], r["skin"], r["paint_kit_name"],
            r["wear_min"], r["wear_max"],
        ])

    for col, width in zip(["A","B","C","D","E","F","G","H"],
                           [12,  12,  52,  22,  30,  40,  12,  12]):
        ws.column_dimensions[col].width = width

    wb.save(path)
    print(f"  {path.name} ({path.stat().st_size // 1024}K)")


def _verify(all_skins: list[dict]):
    checks = {
        "AK-47 | Asiimov":               (7,   801),
        "AWP | Dragon Lore":             (9,   344),
        "\u2605 Karambit | Marble Fade": (507, 413),
        "\u2605 Bayonet | Doppler":      (500, 418),
    }
    for expected_name, (exp_def, exp_paint) in checks.items():
        found = next((s for s in all_skins if s["defindex"] == exp_def and s["paint_index"] == exp_paint), None)
        status = "OK" if found and found["market_hash_name"] == expected_name else "MISMATCH"
        safe = expected_name.encode("ascii", "replace").decode()
        print(f"  [{status}] def={exp_def:5d} paint={exp_paint:5d} -> {safe}")


if __name__ == "__main__":
    force = "--force" in sys.argv
    main(force_download=force)
